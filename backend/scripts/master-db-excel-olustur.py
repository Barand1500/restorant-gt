#!/usr/bin/env python3
"""MASTER_DB tablo yapisi + MySQL veri tipleri Excel dosyasi olusturur."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Her tablo: (tablo_adi, [(sutun_etiketi, mysql_tipi, db_sutun_adi opsiyonel not)])
TABLOLAR = [
    (
        "moduller",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("modul adi", "VARCHAR(100) — NOT NULL", "modul_adi"),
            ("prefix", "VARCHAR(50) — NOT NULL, UNIQUE"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
            ("guncelleme tarihi", "DATETIME(3) — ON UPDATE", "guncelleme_tarihi"),
            ("durum", "BOOLEAN (TINYINT(1)) — DEFAULT 1"),
        ],
    ),
    (
        "roller",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("rol adi", "VARCHAR(50) — NOT NULL", "rol_adi"),
            ("modul id", "INT — FK → moduller.id, NOT NULL", "modul_id"),
            ("yetki", "JSON — DEFAULT []"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
            ("guncelleme tarihi", "DATETIME(3) — ON UPDATE", "guncelleme_tarihi"),
            ("durum", "BOOLEAN (TINYINT(1)) — DEFAULT 1"),
        ],
    ),
    (
        "bayiler",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("ust id", "INT — FK → bayiler.id, NULL", "ust_id"),
            ("unvan", "VARCHAR(191) — NOT NULL"),
            ("adres", "VARCHAR(500) — NULL"),
            ("il", "VARCHAR(50) — NULL"),
            ("ilce", "VARCHAR(50) — NULL"),
            ("telefon", "VARCHAR(20) — NULL"),
            ("gsm", "VARCHAR(20) — NULL"),
            ("eposta", "VARCHAR(191) — NULL"),
            ("vergi dairesi", "VARCHAR(100) — NULL", "vergi_dairesi"),
            ("vergi no", "VARCHAR(50) — NULL", "vergi_no"),
            ("iskonto", "DECIMAL(5,2) — NULL"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
            ("guncelleme tarihi", "DATETIME(3) — ON UPDATE", "guncelleme_tarihi"),
            ("durum", "BOOLEAN (TINYINT(1)) — DEFAULT 1"),
        ],
    ),
    (
        "firmalar",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("bayi id", "INT — FK → bayiler.id, NOT NULL", "bayi_id"),
            ("tabela adi", "VARCHAR(150) — NULL", "tabela_adi"),
            ("unvan", "VARCHAR(191) — NOT NULL"),
            ("adres", "VARCHAR(500) — NULL"),
            ("il", "VARCHAR(50) — NULL"),
            ("ilce", "VARCHAR(50) — NULL"),
            ("telefon", "VARCHAR(20) — NULL"),
            ("gsm", "VARCHAR(20) — NULL"),
            ("eposta", "VARCHAR(191) — NULL"),
            ("vergi dairesi", "VARCHAR(100) — NULL", "vergi_dairesi"),
            ("vergi no", "VARCHAR(50) — NULL", "vergi_no"),
            ("iskonto", "DECIMAL(5,2) — NULL"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
            ("guncelleme tarihi", "DATETIME(3) — ON UPDATE", "guncelleme_tarihi"),
            ("durum", "BOOLEAN (TINYINT(1)) — DEFAULT 1"),
        ],
    ),
    (
        "subeler",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("firma id", "INT — FK → firmalar.id, NOT NULL", "firma_id"),
            ("sube adi", "VARCHAR(100) — NOT NULL", "sube_adi"),
            ("adres", "VARCHAR(500) — NULL"),
            ("il", "VARCHAR(50) — NULL"),
            ("ilce", "VARCHAR(50) — NULL"),
            ("telefon", "VARCHAR(20) — NULL"),
            ("gsm", "VARCHAR(20) — NULL"),
            ("eposta", "VARCHAR(191) — NULL"),
            ("vergi dairesi", "VARCHAR(100) — NULL", "vergi_dairesi"),
            ("vergi no", "VARCHAR(50) — NULL", "vergi_no"),
            ("sube tipi", "ENUM — restoran|kafe|fast_food|diger, DEFAULT restoran", "sube_tipi"),
            ("db bilgileri", "JSON — NULL", "db_bilgileri"),
            ("lisans bilgileri", "JSON — NULL", "lisans_bilgileri"),
            ("iskonto", "DECIMAL(5,2) — NULL"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
            ("guncelleme tarihi", "DATETIME(3) — ON UPDATE", "guncelleme_tarihi"),
            ("durum", "BOOLEAN (TINYINT(1)) — DEFAULT 1"),
        ],
    ),
    (
        "kullanicilar",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("rol id", "INT — FK → roller.id, NOT NULL", "rol_id"),
            ("bayi id", "INT — FK → bayiler.id, NULL", "bayi_id"),
            ("firma id", "INT — FK → firmalar.id, NULL", "firma_id"),
            ("sube id", "INT — FK → subeler.id, NULL", "sube_id"),
            ("ad soyad", "VARCHAR(150) — NOT NULL", "ad_soyad"),
            ("gsm", "VARCHAR(20) — NULL"),
            ("eposta", "VARCHAR(191) — NOT NULL, UNIQUE"),
            ("sifre", "VARCHAR(255) — NOT NULL (hash)", "sifre_hash"),
            ("kullanici tipi", "ENUM — merkez|bayi|firma|sube, DEFAULT merkez", "kullanici_tipi"),
            ("iskonto", "DECIMAL(5,2) — NULL"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
            ("guncelleme tarihi", "DATETIME(3) — ON UPDATE", "guncelleme_tarihi"),
            ("son giris tarihi", "DATETIME(3) — NULL", "son_giris_tarihi"),
            ("durum", "BOOLEAN (TINYINT(1)) — DEFAULT 1"),
        ],
    ),
    (
        "paketler",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("paket adi", "VARCHAR(100) — NOT NULL", "paket_adi"),
            ("sube sayisi", "INT — DEFAULT 1", "sube_sayisi"),
            ("personel sayisi", "INT — DEFAULT 10", "personel_sayisi"),
            ("masa sayisi", "INT — DEFAULT 50", "masa_sayisi"),
            ("fiyat", "DECIMAL(12,2) — DEFAULT 0"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
            ("guncelleme tarihi", "DATETIME(3) — ON UPDATE", "guncelleme_tarihi"),
            ("durum", "BOOLEAN (TINYINT(1)) — DEFAULT 1"),
        ],
    ),
    (
        "lisanslar",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("firma id", "INT — FK → firmalar.id, NOT NULL", "firma_id"),
            ("paket id", "INT — FK → paketler.id, NOT NULL", "paket_id"),
            ("baslangic tarihi", "DATETIME(3) — DEFAULT now()", "baslangic_tarihi"),
            ("bitis tarihi", "DATETIME(3) — NULL", "bitis_tarihi"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
            ("guncelleme tarihi", "DATETIME(3) — ON UPDATE", "guncelleme"),
            ("durum", "BOOLEAN (TINYINT(1)) — DEFAULT 1"),
        ],
    ),
    (
        "log",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("kullanici id", "INT — FK → kullanicilar.id, NULL", "kullanici_id"),
            ("log tipi", "VARCHAR(30) — NULL", "log_tipi"),
            ("mesaj", "VARCHAR(1000) — NOT NULL"),
            ("ip adresi", "VARCHAR(45) — NULL (IPv6)", "ip_adresi"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
        ],
    ),
    (
        "kullanici_kisayol",
        [
            ("kullanici id", "INT — PK, FK → kullanicilar.id", "kullanici_id"),
            ("harita", "JSON — DEFAULT {}"),
        ],
    ),
    (
        "kullanici_sekme_ayar",
        [
            ("kullanici id", "INT — PK, FK → kullanicilar.id", "kullanici_id"),
            ("ayarlar", "JSON — DEFAULT {}"),
        ],
    ),
    (
        "bildirim",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("sube id", "INT — FK → subeler.id, NULL", "sube_id"),
            ("tip", "ENUM — sistem"),
            ("baslik", "VARCHAR(150) — NOT NULL"),
            ("mesaj", "VARCHAR(1000) — NOT NULL"),
            ("okundu", "BOOLEAN (TINYINT(1)) — DEFAULT 0"),
            ("link", "VARCHAR(500) — NULL"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
        ],
    ),
    (
        "eklenti_kurulum",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("sube id", "INT — DEFAULT 0, UNIQUE(sube_id+eklenti_kodu)", "sube_id"),
            ("eklenti kodu", "VARCHAR(100) — FK → eklenti_katalog.kod", "eklenti_kodu"),
            ("durum", "ENUM — kurulu|aktif|pasif, DEFAULT kurulu"),
            ("kaynak", "ENUM — katalog|yukleme, DEFAULT katalog"),
            ("ayarlar json", "JSON — DEFAULT {}", "ayarlar_json"),
            ("kurulum tarihi", "DATETIME(3) — DEFAULT now()", "kurulum_tarihi"),
            ("guncelleme", "DATETIME(3) — ON UPDATE"),
        ],
    ),
    (
        "eklenti_katalog",
        [
            ("kod", "VARCHAR(100) — PK"),
            ("ad", "VARCHAR(150) — NOT NULL"),
            ("aciklama", "VARCHAR(1000) — NOT NULL"),
            ("gelistirici", "VARCHAR(100) — NOT NULL"),
            ("ikon", "VARCHAR(50) — NOT NULL"),
            ("kategori", "VARCHAR(50) — NOT NULL"),
            ("surum", "VARCHAR(30) — NOT NULL"),
            ("puan", "FLOAT — DEFAULT 0"),
            ("etkin kurulum", "INT — DEFAULT 0", "etkin_kurulum"),
            ("son guncelleme", "DATETIME(3) — DEFAULT now()", "son_guncelleme"),
            ("public hook", "VARCHAR(100) — NULL", "public_hook"),
            ("manifest json", "JSON — NULL", "manifest_json"),
        ],
    ),
    (
        "yedek_kaydi",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("sube id", "INT — FK → subeler.id, NULL", "sube_id"),
            ("kullanici id", "INT — FK → kullanicilar.id, NULL", "kullanici_id"),
            ("kullanici ad", "VARCHAR(150) — NOT NULL", "kullanici_ad"),
            ("kullanici email", "VARCHAR(191) — NOT NULL", "kullanici_email"),
            ("dosya adi", "VARCHAR(255) — NOT NULL", "dosya_adi"),
            ("tip", "ENUM — indir|geri_yukle"),
            ("format", "ENUM — json|sql|rar|zip, DEFAULT json"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
        ],
    ),
    (
        "ayarlar",
        [
            ("id", "INT — PK, AUTO_INCREMENT"),
            ("sube id", "INT — DEFAULT 0 (0=merkez), UNIQUE(sube_id+anahtar)", "sube_id"),
            ("anahtar", "VARCHAR(100) — NOT NULL"),
            ("deger", "JSON — NOT NULL"),
            ("kayit tarihi", "DATETIME(3) — DEFAULT now()", "kayit_tarihi"),
            ("guncelleme tarihi", "DATETIME(3) — ON UPDATE", "guncelleme_tarihi"),
        ],
    ),
]

BASLIK_FILL = PatternFill("solid", fgColor="1F4E79")
BASLIK_FONT = Font(bold=True, color="FFFFFF", size=11)
TABLO_FILL = PatternFill("solid", fgColor="D6E4F0")
TABLO_FONT = Font(bold=True, size=10)
TIP_FONT = Font(size=9, color="333333")
FK_BASLIK_FONT = Font(bold=True, color="1A1A1A", size=11)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Ayni renk = ayni FK iliskisi (hedef tablo.id + ona bagli tum FK sutunlari)
FK_GRUPLARI = [
    (
        "moduller",
        "id",
        [("roller", "modul_id")],
        "BDD7EE",
        "→ moduller.id",
    ),
    (
        "roller",
        "id",
        [("kullanicilar", "rol_id")],
        "E2BFED",
        "→ roller.id",
    ),
    (
        "bayiler",
        "id",
        [
            ("firmalar", "bayi_id"),
            ("kullanicilar", "bayi_id"),
            ("bayiler", "ust_id"),
        ],
        "C6EFCE",
        "→ bayiler.id",
    ),
    (
        "firmalar",
        "id",
        [
            ("subeler", "firma_id"),
            ("kullanicilar", "firma_id"),
            ("lisanslar", "firma_id"),
        ],
        "FFF2CC",
        "→ firmalar.id",
    ),
    (
        "subeler",
        "id",
        [
            ("kullanicilar", "sube_id"),
            ("bildirim", "sube_id"),
            ("yedek_kaydi", "sube_id"),
            ("eklenti_kurulum", "sube_id"),
            ("ayarlar", "sube_id"),
        ],
        "FCE4D6",
        "→ subeler.id",
    ),
    (
        "kullanicilar",
        "id",
        [
            ("log", "kullanici_id"),
            ("kullanici_kisayol", "kullanici_id"),
            ("kullanici_sekme_ayar", "kullanici_id"),
            ("yedek_kaydi", "kullanici_id"),
        ],
        "F4B084",
        "→ kullanicilar.id",
    ),
    (
        "paketler",
        "id",
        [("lisanslar", "paket_id")],
        "D9E1F2",
        "→ paketler.id",
    ),
    (
        "eklenti_katalog",
        "kod",
        [("eklenti_kurulum", "eklenti_kodu")],
        "E2EFDA",
        "→ eklenti_katalog.kod",
    ),
]


def db_sutun_adi(sutun: tuple) -> str:
    if len(sutun) > 2:
        return sutun[2]
    return sutun[0].replace(" ", "_")


def fk_renk_uygula(ws, hucre_haritasi: dict) -> None:
    """FK sutunu ile hedef tablodaki PK sutununu ayni renkle boyar."""
    for hedef_tablo, hedef_sutun, cocuklar, renk, _etiket in FK_GRUPLARI:
        fill = PatternFill("solid", fgColor=renk)

        def boya(tablo: str, sutun: str) -> None:
            anahtar = (tablo, sutun)
            if anahtar not in hucre_haritasi:
                return
            baslik_satir, tip_satir, col = hucre_haritasi[anahtar]
            for satir in (baslik_satir, tip_satir):
                hucre = ws.cell(row=satir, column=col)
                hucre.fill = fill
                if satir == baslik_satir:
                    hucre.font = FK_BASLIK_FONT

        boya(hedef_tablo, hedef_sutun)
        for cocuk_tablo, cocuk_sutun in cocuklar:
            boya(cocuk_tablo, cocuk_sutun)


def fk_legend_ekle(ws, baslangic_satir: int) -> None:
    ws.cell(row=baslangic_satir, column=1, value="FK RENK LEGENDI")
    ws.cell(row=baslangic_satir, column=1).font = Font(bold=True, size=11)
    satir = baslangic_satir + 1
    for _hedef_tablo, _hedef_sutun, cocuklar, renk, etiket in FK_GRUPLARI:
        fill = PatternFill("solid", fgColor=renk)
        ws.cell(row=satir, column=1).fill = fill
        ws.cell(row=satir, column=1).border = BORDER
        cocuk_metin = ", ".join(f"{t}.{s}" for t, s in cocuklar)
        ws.cell(
            row=satir,
            column=2,
            value=f"{etiket}  —  FK: {cocuk_metin}",
        )
        ws.cell(row=satir, column=2).font = Font(size=9)
        satir += 1


def olustur(cikti: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER DB"

    ws["A1"] = "Guzel Teknoloji — MASTER DB (MySQL / Prisma)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    ws["A2"] = "Kaynak: backend/prisma/schema.prisma"
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    row = 4
    hucre_haritasi: dict[tuple[str, str], tuple[int, int, int]] = {}

    for tablo_adi, sutunlar in TABLOLAR:
        ws.cell(row=row, column=1, value=tablo_adi.upper())
        ws.cell(row=row, column=1).fill = TABLO_FILL
        ws.cell(row=row, column=1).font = TABLO_FONT
        ws.cell(row=row, column=1).border = BORDER

        for col_idx, sutun in enumerate(sutunlar, start=2):
            etiket = sutun[0]
            tip = sutun[1]
            db_adi = db_sutun_adi(sutun)

            hucre_haritasi[(tablo_adi, db_adi)] = (row, row + 1, col_idx)

            c1 = ws.cell(row=row, column=col_idx, value=etiket)
            c1.fill = BASLIK_FILL
            c1.font = BASLIK_FONT
            c1.alignment = Alignment(horizontal="center", wrap_text=True)
            c1.border = BORDER

            tip_metin = tip
            if db_adi.replace(" ", "_") != etiket.replace(" ", "_"):
                tip_metin = f"{tip}\n[{db_adi}]"

            c2 = ws.cell(row=row + 1, column=col_idx, value=tip_metin)
            c2.font = TIP_FONT
            c2.alignment = Alignment(vertical="top", wrap_text=True)
            c2.border = BORDER

        ws.cell(row=row + 1, column=1, value="veri tipi")
        ws.cell(row=row + 1, column=1).font = Font(bold=True, size=9)
        ws.cell(row=row + 1, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
        ws.cell(row=row + 1, column=1).border = BORDER

        row += 3

    fk_renk_uygula(ws, hucre_haritasi)
    fk_legend_ekle(ws, row + 1)

    # Ozet sayfasi
    ozet = wb.create_sheet("Ozet Liste")
    ozet.append(["Tablo", "Sutun (Excel)", "DB sutun adi", "MySQL tipi", "Not", "FK Grubu"])
    for cell in ozet[1]:
        cell.fill = BASLIK_FILL
        cell.font = BASLIK_FONT

    fk_lookup: dict[tuple[str, str], str] = {}
    for hedef_tablo, hedef_sutun, cocuklar, _renk, etiket in FK_GRUPLARI:
        fk_lookup[(hedef_tablo, hedef_sutun)] = etiket
        for cocuk_tablo, cocuk_sutun in cocuklar:
            fk_lookup[(cocuk_tablo, cocuk_sutun)] = etiket

    for tablo_adi, sutunlar in TABLOLAR:
        for sutun in sutunlar:
            etiket = sutun[0]
            tip = sutun[1]
            db_adi = db_sutun_adi(sutun)
            fk_grubu = fk_lookup.get((tablo_adi, db_adi), "")
            ozet.append([tablo_adi, etiket, db_adi, tip.split(" — ")[0], tip, fk_grubu])

    for col in ozet.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ozet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    ws.column_dimensions["A"].width = 22
    for col_letter in "BCDEFGHIJKLMNOPQRSTUVWXYZ":
        ws.column_dimensions[col_letter].width = 18

    wb.save(cikti)
    print(f"Olusturuldu: {cikti}")


if __name__ == "__main__":
    docs = Path(__file__).resolve().parent.parent / "docs"
    docs.mkdir(parents=True, exist_ok=True)
    out = docs / "MASTER_DB_veri_tipleri.xlsx"
    try:
        olustur(out)
    except PermissionError:
        yedek = docs / "MASTER_DB_veri_tipleri_fk.xlsx"
        olustur(yedek)
        print(f"Ana dosya acik — yedek olusturuldu: {yedek}")
